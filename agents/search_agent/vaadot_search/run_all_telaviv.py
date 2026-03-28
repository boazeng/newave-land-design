#!/usr/bin/env python3
"""
סקריפט ראשי להרצת כל הסקריפטים של ועדות מחוז תל אביב.
מוריד פרוטוקולים מכל 11 הועדות ומייצר טבלת סיכום.

שלב 1: מציאת Site IDs של ועדות Complot (אם לא ידוע)
שלב 2: הרצת סקריפטים לכל ועדה
שלב 3: יצירת טבלת סיכום Excel

התקנה:
    pip install requests lxml beautifulsoup4 openpyxl

הרצה:
    python3 run_all_telaviv.py                    # הרצה מלאה - חילוץ + הורדה
    python3 run_all_telaviv.py --count-only       # רק ספירה (מהיר)
    python3 run_all_telaviv.py --no-download      # חילוץ קישורים בלבד
    python3 run_all_telaviv.py --skip-complot-scan # דילוג על סריקת IDs

פרמטרים:
    --start-year        שנה ראשונה (ברירת מחדל: 2015)
    --end-year          שנה אחרונה (ברירת מחדל: 2026)
    --output-dir        תיקיית פלט (ברירת מחדל: output)
    --count-only        רק ספירת פרוטוקולים (בלי הורדה)
    --no-download       חילוץ קישורים בלבד
    --skip-complot-scan דילוג על סריקת Complot IDs (משתמש בברירות מחדל)
"""

import os
import sys
import json
import time
import argparse
import subprocess
from datetime import datetime
from pathlib import Path

# ==============================================================================
# COMMITTEE CONFIGURATION - Tel Aviv District
# ==============================================================================

COMMITTEES = [
    # Complot-based committees (API: GetMeetingByDate + GetMeetingDocs)
    {
        'name': 'רמת גן',
        'name_en': 'ramat_gan',
        'platform': 'complot',
        'site_id': 3,           # CONFIRMED
        'city_name': 'רמת_גן',
    },
    {
        'name': 'חולון',
        'name_en': 'holon',
        'platform': 'complot',
        'site_id': 34,          # CONFIRMED
        'city_name': 'חולון',
    },
    {
        'name': 'בת ים',
        'name_en': 'bat_yam',
        'platform': 'complot',
        'site_id': 81,          # CONFIRMED
        'city_name': 'בת_ים',
    },
    {
        'name': 'אור יהודה',
        'name_en': 'or_yehuda',
        'platform': 'complot',
        'site_id': 73,          # CONFIRMED
        'city_name': 'אור_יהודה',
    },
    {
        'name': 'רמת השרון',
        'name_en': 'ramat_hasharon',
        'platform': 'complot',
        'site_id': 118,         # CONFIRMED
        'city_name': 'רמת_השרון',
    },
    {
        'name': 'גבעתיים',
        'name_en': 'givatayim',
        'platform': 'complot',
        'site_id': 98,          # CONFIRMED
        'city_name': 'גבעתיים',
    },
    {
        'name': 'הרצליה',
        'name_en': 'herzliya',
        'platform': 'complot',
        'site_id': 121,         # CONFIRMED
        'city_name': 'הרצליה',
    },
    {
        'name': 'בני ברק',
        'name_en': 'bnei_brak',
        'platform': 'complot',
        'site_id': 75,          # CONFIRMED
        'city_name': 'בני_ברק',
    },

    # Bartech-based committees
    {
        'name': 'קרית אונו',
        'name_en': 'kiryat_ono',
        'platform': 'bartech',
        'city_code': 'ono',
        'city_name': 'קרית_אונו',
    },
    {
        'name': 'אזור',
        'name_en': 'azor',
        'platform': 'bartech',
        'city_code': 'azr',
        'city_name': 'אזור',
    },

    # Tel Aviv - independent site
    {
        'name': 'תל אביב-יפו',
        'name_en': 'tel_aviv',
        'platform': 'telaviv',
        'city_name': 'תל_אביב_יפו',
    },
]


def find_complot_site_ids(output_dir):
    """Run the site ID finder and return the mapping."""
    script = os.path.join(os.path.dirname(__file__), 'find_complot_ids.py')
    if not os.path.exists(script):
        print("ERROR: find_complot_ids.py not found!")
        return {}

    print("\n" + "=" * 60)
    print("STEP 1: Scanning Complot Site IDs")
    print("=" * 60)

    result = subprocess.run(
        [sys.executable, script],
        capture_output=True, text=True, timeout=600,
        cwd=output_dir
    )
    print(result.stdout)
    if result.stderr:
        print("STDERR:", result.stderr[:500])

    # Read results
    json_path = os.path.join(output_dir, 'complot_site_ids.json')
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def match_complot_ids(committees, site_id_map):
    """Match discovered site IDs to our committees."""
    for committee in committees:
        if committee['platform'] != 'complot' or committee.get('site_id'):
            continue

        subdomain = committee.get('complot_subdomain', '').lower()
        city_name = committee.get('city_name', '').lower()

        for sid, info in site_id_map.items():
            city = info.get('city', '').lower().replace('-', '').replace('_', '').replace(' ', '')
            if subdomain and subdomain in city:
                committee['site_id'] = int(sid)
                print(f"  Matched {committee['name']}: Site ID = {sid} (city: {info['city']})")
                break
            elif city_name.replace('_', '') in city:
                committee['site_id'] = int(sid)
                print(f"  Matched {committee['name']}: Site ID = {sid} (city: {info['city']})")
                break

    return committees


def run_complot_scraper(committee, args):
    """Run complot_scraper.py for a single committee."""
    if not committee.get('site_id'):
        print(f"  SKIP {committee['name']}: No site ID found")
        return None

    script = os.path.join(os.path.dirname(__file__), 'complot_scraper.py')
    cmd = [
        sys.executable, script,
        '--site-id', str(committee['site_id']),
        '--city-name', committee['city_name'],
        '--start-year', str(args.start_year),
        '--end-year', str(args.end_year),
        '--output-dir', args.output_dir,
        '--protocols-only',
    ]
    if args.count_only:
        cmd.append('--count-only')
    elif args.no_download:
        cmd.append('--no-download')

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=1800)
    print(result.stdout[-500:] if len(result.stdout) > 500 else result.stdout)
    if result.stderr:
        print("  STDERR:", result.stderr[:300])
    return result.returncode


def run_bartech_scraper(committee, args):
    """Run bartech_scraper.py for a single committee."""
    script = os.path.join(os.path.dirname(__file__), 'bartech_scraper.py')
    cmd = [
        sys.executable, script,
        '--city-code', committee['city_code'],
        '--city-name', committee['city_name'],
        '--start-year', str(args.start_year),
        '--end-year', str(args.end_year),
        '--output-dir', args.output_dir,
        '--protocols-only',
    ]
    if args.count_only:
        cmd.append('--count-only')
    elif args.no_download:
        cmd.append('--no-download')

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=1800)
    print(result.stdout[-500:] if len(result.stdout) > 500 else result.stdout)
    if result.stderr:
        print("  STDERR:", result.stderr[:300])
    return result.returncode


def run_sharepoint_scraper(committee, args):
    """Run sharepoint_scraper.py for a single committee."""
    script = os.path.join(os.path.dirname(__file__), 'sharepoint_scraper.py')
    cmd = [
        sys.executable, script,
        '--city', committee['sp_city'],
        '--start-year', str(args.start_year),
        '--end-year', str(args.end_year),
        '--output-dir', args.output_dir,
    ]
    if args.count_only:
        cmd.append('--count-only')
    elif args.no_download:
        cmd.append('--no-download')

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=1800)
    print(result.stdout[-500:] if len(result.stdout) > 500 else result.stdout)
    if result.stderr:
        print("  STDERR:", result.stderr[:300])
    return result.returncode


def run_telaviv_scraper(committee, args):
    """Run telaviv_scraper.py."""
    script = os.path.join(os.path.dirname(__file__), 'telaviv_scraper.py')
    cmd = [
        sys.executable, script,
        '--start-year', str(args.start_year),
        '--end-year', str(args.end_year),
        '--output-dir', args.output_dir,
    ]
    if args.count_only:
        cmd.append('--count-only')
    elif args.no_download:
        cmd.append('--no-download')

    result = subprocess.run(cmd, capture_output=True, text=True, timeout=1800)
    print(result.stdout[-500:] if len(result.stdout) > 500 else result.stdout)
    if result.stderr:
        print("  STDERR:", result.stderr[:300])
    return result.returncode


def collect_results(committees, output_dir):
    """Collect protocol counts from all JSON result files."""
    results = {}

    for committee in committees:
        city_name = committee['city_name']
        json_path = os.path.join(output_dir, city_name, f'{city_name}_data.json')

        if not os.path.exists(json_path):
            results[committee['name']] = {'counts': {}, 'total': 0, 'status': 'NO DATA'}
            continue

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            counts = data.get('counts_by_year', {})
            # Convert string keys to int
            counts = {int(k): v for k, v in counts.items()}
            total = data.get('total_documents', sum(counts.values()))
            results[committee['name']] = {
                'counts': counts,
                'total': total,
                'status': 'OK'
            }
        except Exception as e:
            results[committee['name']] = {'counts': {}, 'total': 0, 'status': f'ERROR: {e}'}

    return results


def generate_summary_excel(results, start_year, end_year, output_dir):
    """Generate summary Excel file with protocol counts."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: openpyxl not installed. pip install openpyxl")
        # Fallback to CSV
        generate_summary_csv(results, start_year, end_year, output_dir)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "סיכום פרוטוקולים"
    ws.sheet_view.rightToLeft = True

    # Styles
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    total_fill = PatternFill('solid', fgColor='D6E4F0')
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center')

    years = list(range(start_year, end_year + 1))

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(years) + 3)
    title_cell = ws.cell(row=1, column=1,
                         value=f'פרוטוקולי ועדות תכנון ובניה - מחוז תל אביב ({start_year}-{end_year})')
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # Headers
    headers = ['ועדה', 'פלטפורמה'] + [str(y) for y in years] + ['סה"כ', 'סטטוס']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Data rows
    row = 4
    committee_names = list(results.keys())
    platform_map = {c['name']: c['platform'] for c in COMMITTEES}

    for name in committee_names:
        data = results[name]
        counts = data['counts']
        total = data['total']
        status = data['status']
        platform = platform_map.get(name, '?')

        ws.cell(row=row, column=1, value=name).border = border
        ws.cell(row=row, column=2, value=platform).border = border

        row_total = 0
        for col_idx, year in enumerate(years):
            count = counts.get(year, 0)
            cell = ws.cell(row=row, column=col_idx + 3, value=count if count > 0 else '')
            cell.alignment = center
            cell.border = border
            row_total += count

        total_cell = ws.cell(row=row, column=len(years) + 3, value=row_total)
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.alignment = center
        total_cell.border = border

        status_cell = ws.cell(row=row, column=len(years) + 4, value=status)
        status_cell.border = border
        if status != 'OK':
            status_cell.font = Font(color='FF0000')

        row += 1

    # Total row
    ws.cell(row=row, column=1, value='סה"כ').font = total_font
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2).fill = total_fill
    ws.cell(row=row, column=2).border = border

    for col_idx, year in enumerate(years):
        col = col_idx + 3
        total = sum(results[name]['counts'].get(year, 0) for name in committee_names)
        cell = ws.cell(row=row, column=col, value=total if total > 0 else '')
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center
        cell.border = border

    grand_total = sum(data['total'] for data in results.values())
    gt_cell = ws.cell(row=row, column=len(years) + 3, value=grand_total)
    gt_cell.font = Font(bold=True, size=12)
    gt_cell.fill = total_fill
    gt_cell.alignment = center
    gt_cell.border = border

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    for i in range(len(years)):
        col_letter = chr(ord('C') + i) if i < 24 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = 8

    # Save
    excel_path = os.path.join(output_dir, 'סיכום_פרוטוקולים_מחוז_תל_אביב.xlsx')
    wb.save(excel_path)
    print(f"\n  Summary saved to: {excel_path}")
    return excel_path


def generate_summary_csv(results, start_year, end_year, output_dir):
    """Fallback: generate CSV summary."""
    import csv
    years = list(range(start_year, end_year + 1))
    csv_path = os.path.join(output_dir, 'summary_telaviv.csv')

    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['ועדה', 'פלטפורמה'] + [str(y) for y in years] + ['סהכ', 'סטטוס'])

        platform_map = {c['name']: c['platform'] for c in COMMITTEES}
        for name, data in results.items():
            row = [name, platform_map.get(name, '?')]
            for year in years:
                row.append(data['counts'].get(year, 0))
            row.append(data['total'])
            row.append(data['status'])
            writer.writerow(row)

    print(f"\n  Summary CSV saved to: {csv_path}")


def main():
    parser = argparse.ArgumentParser(description='Tel Aviv District - All Committees Protocol Scraper')
    parser.add_argument('--start-year', type=int, default=2015)
    parser.add_argument('--end-year', type=int, default=2026)
    parser.add_argument('--output-dir', type=str, default='output')
    parser.add_argument('--count-only', action='store_true', help='Only count protocols')
    parser.add_argument('--no-download', action='store_true', help='Extract links only')
    parser.add_argument('--skip-complot-scan', action='store_true', help='Skip Complot ID scan')
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    print("=" * 70)
    print("  מחוז תל אביב - הורדת פרוטוקולי ועדות תכנון ובניה")
    print(f"  שנים: {args.start_year}-{args.end_year}")
    print(f"  פלט: {args.output_dir}")
    print(f"  מצב: {'ספירה בלבד' if args.count_only else 'ללא הורדה' if args.no_download else 'הורדה מלאה'}")
    print(f"  ועדות: {len(COMMITTEES)}")
    print("=" * 70)

    committees = COMMITTEES.copy()

    # Step 1: Find Complot IDs if needed
    need_ids = any(c['platform'] == 'complot' and not c.get('site_id') for c in committees)
    if need_ids and not args.skip_complot_scan:
        site_ids = find_complot_site_ids(args.output_dir)
        if site_ids:
            committees = match_complot_ids(committees, site_ids)
        else:
            print("  WARNING: Could not discover Complot site IDs.")
            print("  Run find_complot_ids.py manually and update COMMITTEES config.")
    elif need_ids:
        print("\n  WARNING: Some Complot committees have no site ID. Run find_complot_ids.py first.")
        print("  Committees without IDs will be skipped.")

    # Step 2: Run scrapers for each committee
    print("\n" + "=" * 70)
    print("  STEP 2: Running scrapers for each committee")
    print("=" * 70)

    runners = {
        'complot': run_complot_scraper,
        'bartech': run_bartech_scraper,
        'sharepoint': run_sharepoint_scraper,
        'telaviv': run_telaviv_scraper,
    }

    for i, committee in enumerate(committees):
        print(f"\n{'─' * 60}")
        print(f"  [{i+1}/{len(committees)}] {committee['name']} ({committee['platform']})")
        print(f"{'─' * 60}")

        runner = runners.get(committee['platform'])
        if runner:
            try:
                runner(committee, args)
            except subprocess.TimeoutExpired:
                print(f"  TIMEOUT for {committee['name']}")
            except Exception as e:
                print(f"  ERROR for {committee['name']}: {e}")
        else:
            print(f"  SKIP: Unknown platform '{committee['platform']}'")

    # Step 3: Collect results and generate summary
    print("\n" + "=" * 70)
    print("  STEP 3: Generating summary")
    print("=" * 70)

    results = collect_results(committees, args.output_dir)
    generate_summary_excel(results, args.start_year, args.end_year, args.output_dir)

    # Print final summary
    print("\n" + "=" * 70)
    print("  FINAL SUMMARY")
    print("=" * 70)
    for name, data in results.items():
        status = '✅' if data['status'] == 'OK' else '❌'
        print(f"  {status} {name}: {data['total']} protocols ({data['status']})")

    total = sum(d['total'] for d in results.values())
    print(f"\n  Total protocols across all committees: {total}")
    print("=" * 70)


if __name__ == '__main__':
    main()
